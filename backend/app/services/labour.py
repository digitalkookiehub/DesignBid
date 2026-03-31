import logging
from decimal import Decimal
from io import BytesIO

from sqlalchemy.orm import Session

from app.exceptions import BadRequestError, ConflictError, NotFoundError
from app.models.labour import AssignmentStatus, Labour, LabourSpecialization, ProjectLabourAssignment
from app.models.project import Project
from app.schemas.labour import AssignLabourRequest, LabourCreate, LabourUpdate, UpdateAssignmentRequest

logger = logging.getLogger(__name__)


# --- Labour Directory ---

def get_labours(db: Session, user_id: int, specialization: str = "", search: str = "", active_only: bool = True) -> list[dict]:
    query = db.query(Labour).filter(Labour.user_id == user_id)
    if active_only:
        query = query.filter(Labour.is_active == True)  # noqa: E712
    if specialization:
        query = query.filter(Labour.specialization == LabourSpecialization(specialization))
    if search:
        query = query.filter(
            (Labour.name.ilike(f"%{search}%")) | (Labour.phone.ilike(f"%{search}%"))
        )
    labours = query.order_by(Labour.name).all()

    # Enrich with current project info
    result = []
    for labour in labours:
        current_assignment = (
            db.query(ProjectLabourAssignment)
            .join(Project)
            .filter(
                ProjectLabourAssignment.labour_id == labour.id,
                ProjectLabourAssignment.status.in_([AssignmentStatus.assigned, AssignmentStatus.working]),
            )
            .first()
        )
        current_project = None
        if current_assignment:
            project = db.query(Project).filter(Project.id == current_assignment.project_id).first()
            current_project = project.name if project else None

        labour_dict = {
            "id": labour.id,
            "user_id": labour.user_id,
            "name": labour.name,
            "phone": labour.phone,
            "alt_phone": labour.alt_phone,
            "email": labour.email,
            "specialization": labour.specialization.value if hasattr(labour.specialization, "value") else str(labour.specialization),
            "daily_rate": float(labour.daily_rate) if labour.daily_rate else None,
            "address": labour.address,
            "city": labour.city,
            "id_proof_type": labour.id_proof_type,
            "id_proof_number": labour.id_proof_number,
            "experience_years": labour.experience_years,
            "rating": labour.rating,
            "notes": labour.notes,
            "is_active": labour.is_active,
            "current_project": current_project,
            "created_at": labour.created_at.isoformat() if labour.created_at else None,
            "updated_at": labour.updated_at.isoformat() if labour.updated_at else None,
        }
        result.append(labour_dict)
    return result


def get_labour(db: Session, user_id: int, labour_id: int) -> Labour:
    labour = db.query(Labour).filter(Labour.id == labour_id, Labour.user_id == user_id).first()
    if not labour:
        raise NotFoundError("Labour")
    return labour


def create_labour(db: Session, user_id: int, data: LabourCreate) -> Labour:
    labour = Labour(
        user_id=user_id,
        name=data.name,
        phone=data.phone,
        alt_phone=data.alt_phone,
        email=data.email,
        specialization=data.specialization,
        daily_rate=Decimal(str(data.daily_rate)) if data.daily_rate else None,
        address=data.address,
        city=data.city,
        id_proof_type=data.id_proof_type,
        id_proof_number=data.id_proof_number,
        experience_years=data.experience_years,
        notes=data.notes,
    )
    db.add(labour)
    db.commit()
    db.refresh(labour)
    logger.info("Labour created: %s (id=%d)", labour.name, labour.id)
    return labour


# --- Bulk Upload / Export ---

VALID_SPECIALIZATIONS = {s.value for s in LabourSpecialization}

TEMPLATE_HEADERS = ["Name*", "Phone*", "Specialization", "Daily Rate", "City", "Email", "Alt Phone", "Experience (Years)", "Address", "ID Proof Type", "ID Proof Number", "Notes"]


def generate_template() -> BytesIO:
    """Generate an Excel template for bulk labour upload."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Labour Data"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")

    for col_idx, header in enumerate(TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

    # Add sample row
    sample = ["Ramesh Kumar", "9876543210", "carpentry", "800", "Bangalore", "", "", "10", "", "Aadhaar", "", ""]
    for col_idx, val in enumerate(sample, 1):
        ws.cell(row=2, column=col_idx, value=val)

    # Add specialization reference in a second sheet
    ws2 = wb.create_sheet("Specializations")
    ws2.cell(row=1, column=1, value="Valid Specialization Values").font = Font(bold=True)
    for i, s in enumerate(sorted(VALID_SPECIALIZATIONS), 2):
        ws2.cell(row=i, column=1, value=s)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def bulk_upload(db: Session, user_id: int, file_bytes: bytes) -> dict:
    """Parse Excel file and create labourers. Returns summary."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes))
    except Exception:
        raise BadRequestError("Invalid Excel file. Please use .xlsx format.")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    created = 0
    skipped = 0
    errors: list[str] = []

    for row_num, row in enumerate(rows, 2):
        if not row or not row[0]:
            continue

        name = str(row[0]).strip()
        phone = str(row[1]).strip() if row[1] else ""

        if not name or not phone:
            errors.append(f"Row {row_num}: Name and Phone are required")
            skipped += 1
            continue

        # Check duplicate phone
        existing = db.query(Labour).filter(Labour.user_id == user_id, Labour.phone == phone, Labour.is_active == True).first()  # noqa: E712
        if existing:
            errors.append(f"Row {row_num}: Phone {phone} already exists ({existing.name})")
            skipped += 1
            continue

        specialization = str(row[2]).strip().lower() if row[2] else "other"
        if specialization not in VALID_SPECIALIZATIONS:
            specialization = "other"

        daily_rate = None
        if row[3]:
            try:
                daily_rate = Decimal(str(row[3]))
            except Exception:
                pass

        experience = None
        if row[7]:
            try:
                experience = int(row[7])
            except Exception:
                pass

        labour = Labour(
            user_id=user_id,
            name=name,
            phone=phone,
            specialization=specialization,
            daily_rate=daily_rate,
            city=str(row[4]).strip() if row[4] else None,
            email=str(row[5]).strip() if row[5] else None,
            alt_phone=str(row[6]).strip() if row[6] else None,
            experience_years=experience,
            address=str(row[8]).strip() if row[8] else None,
            id_proof_type=str(row[9]).strip() if row[9] else None,
            id_proof_number=str(row[10]).strip() if len(row) > 10 and row[10] else None,
            notes=str(row[11]).strip() if len(row) > 11 and row[11] else None,
        )
        db.add(labour)
        created += 1

    db.commit()
    logger.info("Bulk upload: %d created, %d skipped for user %d", created, skipped, user_id)

    return {
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "total_rows": len(rows),
    }


def export_labours(db: Session, user_id: int) -> BytesIO:
    """Export all labourers as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    labours = db.query(Labour).filter(Labour.user_id == user_id, Labour.is_active == True).order_by(Labour.name).all()  # noqa: E712

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Labour Data"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")

    headers = ["Name", "Phone", "Specialization", "Daily Rate", "City", "Email", "Alt Phone", "Experience", "Address", "ID Proof Type", "ID Proof Number", "Rating", "Notes"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

    for row_idx, l in enumerate(labours, 2):
        ws.cell(row=row_idx, column=1, value=l.name)
        ws.cell(row=row_idx, column=2, value=l.phone)
        ws.cell(row=row_idx, column=3, value=l.specialization.value if hasattr(l.specialization, "value") else str(l.specialization))
        ws.cell(row=row_idx, column=4, value=float(l.daily_rate) if l.daily_rate else None)
        ws.cell(row=row_idx, column=5, value=l.city)
        ws.cell(row=row_idx, column=6, value=l.email)
        ws.cell(row=row_idx, column=7, value=l.alt_phone)
        ws.cell(row=row_idx, column=8, value=l.experience_years)
        ws.cell(row=row_idx, column=9, value=l.address)
        ws.cell(row=row_idx, column=10, value=l.id_proof_type)
        ws.cell(row=row_idx, column=11, value=l.id_proof_number)
        ws.cell(row=row_idx, column=12, value=l.rating)
        ws.cell(row=row_idx, column=13, value=l.notes)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def update_labour(db: Session, user_id: int, labour_id: int, data: LabourUpdate) -> Labour:
    labour = get_labour(db, user_id, labour_id)
    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        if key == "daily_rate" and value is not None:
            value = Decimal(str(value))
        setattr(labour, key, value)
    db.commit()
    db.refresh(labour)
    logger.info("Labour updated: id=%d", labour_id)
    return labour


def delete_labour(db: Session, user_id: int, labour_id: int) -> None:
    labour = get_labour(db, user_id, labour_id)
    labour.is_active = False
    db.commit()
    logger.info("Labour deactivated: id=%d", labour_id)


# --- Project Assignments ---

def get_project_labours(db: Session, user_id: int, project_id: int) -> list[dict]:
    """Get all labourers assigned to a project."""
    # Verify project ownership
    project = db.query(Project).filter(Project.id == project_id, Project.user_id == user_id).first()
    if not project:
        raise NotFoundError("Project")

    assignments = (
        db.query(ProjectLabourAssignment)
        .filter(ProjectLabourAssignment.project_id == project_id)
        .order_by(ProjectLabourAssignment.created_at)
        .all()
    )

    result = []
    for a in assignments:
        labour = db.query(Labour).filter(Labour.id == a.labour_id).first()
        result.append({
            "id": a.id,
            "project_id": a.project_id,
            "labour_id": a.labour_id,
            "labour_name": labour.name if labour else None,
            "labour_phone": labour.phone if labour else None,
            "labour_specialization": (labour.specialization.value if labour and hasattr(labour.specialization, "value") else None),
            "role": a.role,
            "daily_rate": float(a.daily_rate) if a.daily_rate else None,
            "start_date": a.start_date.isoformat() if a.start_date else None,
            "end_date": a.end_date.isoformat() if a.end_date else None,
            "status": a.status.value if hasattr(a.status, "value") else str(a.status),
            "notes": a.notes,
            "created_at": a.created_at.isoformat() if a.created_at else None,
            "updated_at": a.updated_at.isoformat() if a.updated_at else None,
        })
    return result


def assign_labour(db: Session, user_id: int, project_id: int, data: AssignLabourRequest) -> ProjectLabourAssignment:
    """Assign a labourer to a project."""
    project = db.query(Project).filter(Project.id == project_id, Project.user_id == user_id).first()
    if not project:
        raise NotFoundError("Project")

    labour = get_labour(db, user_id, data.labour_id)

    existing = db.query(ProjectLabourAssignment).filter(
        ProjectLabourAssignment.project_id == project_id,
        ProjectLabourAssignment.labour_id == data.labour_id,
    ).first()
    if existing:
        raise ConflictError(f"{labour.name} is already assigned to this project")

    assignment = ProjectLabourAssignment(
        project_id=project_id,
        labour_id=data.labour_id,
        role=data.role,
        daily_rate=Decimal(str(data.daily_rate)) if data.daily_rate else labour.daily_rate,
        start_date=data.start_date,
        end_date=data.end_date,
        notes=data.notes,
    )
    db.add(assignment)
    db.commit()
    db.refresh(assignment)
    logger.info("Labour %s assigned to project %d", labour.name, project_id)
    return assignment


def update_assignment(db: Session, user_id: int, project_id: int, assignment_id: int, data: UpdateAssignmentRequest) -> ProjectLabourAssignment:
    project = db.query(Project).filter(Project.id == project_id, Project.user_id == user_id).first()
    if not project:
        raise NotFoundError("Project")

    assignment = db.query(ProjectLabourAssignment).filter(
        ProjectLabourAssignment.id == assignment_id,
        ProjectLabourAssignment.project_id == project_id,
    ).first()
    if not assignment:
        raise NotFoundError("Assignment")

    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        if key == "daily_rate" and value is not None:
            value = Decimal(str(value))
        if key == "status" and value is not None:
            value = AssignmentStatus(value)
        setattr(assignment, key, value)
    db.commit()
    db.refresh(assignment)
    return assignment


def remove_assignment(db: Session, user_id: int, project_id: int, assignment_id: int) -> None:
    project = db.query(Project).filter(Project.id == project_id, Project.user_id == user_id).first()
    if not project:
        raise NotFoundError("Project")

    assignment = db.query(ProjectLabourAssignment).filter(
        ProjectLabourAssignment.id == assignment_id,
        ProjectLabourAssignment.project_id == project_id,
    ).first()
    if not assignment:
        raise NotFoundError("Assignment")

    db.delete(assignment)
    db.commit()
    logger.info("Labour removed from project %d", project_id)
