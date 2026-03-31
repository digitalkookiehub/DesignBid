import logging
from decimal import Decimal
from io import BytesIO

from sqlalchemy.orm import Session, joinedload

from app.exceptions import NotFoundError
from app.models.project import Project
from app.models.quotation import Quotation, QuotationSection

logger = logging.getLogger(__name__)


def _format_inr(amount: Decimal | float) -> str:
    """Format number as Indian Rupees."""
    val = float(amount)
    if val >= 10000000:
        return f"Rs.{val / 10000000:,.2f} Cr"
    if val >= 100000:
        return f"Rs.{val / 100000:,.2f} L"
    return f"Rs.{val:,.2f}"


def export_quotation_xlsx(db: Session, user_id: int, quotation_id: int) -> BytesIO:
    """Export quotation as Excel spreadsheet."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        raise NotFoundError("openpyxl not installed. Run: pip install openpyxl")

    quotation = (
        db.query(Quotation)
        .options(joinedload(Quotation.sections).joinedload(QuotationSection.line_items))
        .join(Project)
        .filter(Quotation.id == quotation_id, Project.user_id == user_id)
        .first()
    )
    if not quotation:
        raise NotFoundError("Quotation")

    project = db.query(Project).filter(Project.id == quotation.project_id).first()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotation"

    # Styles
    header_font = Font(bold=True, size=14, color="1F4E79")
    section_font = Font(bold=True, size=11, color="1F4E79")
    section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    col_header_font = Font(bold=True, size=10, color="FFFFFF")
    col_header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    total_font = Font(bold=True, size=11)
    grand_total_font = Font(bold=True, size=13, color="1F4E79")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = f"QUOTATION - {project.name if project else 'Project'}"
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Version {quotation.version} | Date: {quotation.created_at.strftime('%d-%b-%Y')}"
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4

    # Column headers
    headers = ["S.No", "Room", "Item", "Qty", "Unit", "Rate (Rs.)", "Amount (Rs.)"]
    col_widths = [6, 18, 30, 10, 8, 14, 16]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    row += 1
    serial = 1

    for section in sorted(quotation.sections, key=lambda s: s.sort_order):
        # Section header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=section.section_name)
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
        total_cell = ws.cell(row=row, column=7, value=float(section.section_total))
        total_cell.font = section_font
        total_cell.fill = section_fill
        total_cell.number_format = '#,##0.00'
        total_cell.border = thin_border
        row += 1

        for item in sorted(section.line_items, key=lambda i: i.sort_order):
            ws.cell(row=row, column=1, value=serial).border = thin_border
            ws.cell(row=row, column=2, value=item.room_name or "-").border = thin_border
            ws.cell(row=row, column=3, value=item.item_name).border = thin_border
            qty_cell = ws.cell(row=row, column=4, value=float(item.quantity))
            qty_cell.number_format = '#,##0.00'
            qty_cell.border = thin_border
            ws.cell(row=row, column=5, value=item.unit).border = thin_border
            rate_cell = ws.cell(row=row, column=6, value=float(item.rate))
            rate_cell.number_format = '#,##0.00'
            rate_cell.border = thin_border
            amt_cell = ws.cell(row=row, column=7, value=float(item.amount))
            amt_cell.number_format = '#,##0.00'
            amt_cell.border = thin_border
            serial += 1
            row += 1

        row += 1  # Blank row between sections

    # Totals
    row += 1
    for label, value in [
        ("Subtotal", float(quotation.subtotal)),
        (f"Discount", float(quotation.discount_amount)),
        ("Taxable Amount", float(quotation.taxable_amount)),
        (f"GST ({quotation.tax_rate}%)", float(quotation.tax_amount)),
    ]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=label).font = total_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
        cell = ws.cell(row=row, column=7, value=value)
        cell.font = total_font
        cell.number_format = '#,##0.00'
        cell.border = thin_border
        row += 1

    # Grand total
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="GRAND TOTAL").font = grand_total_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    gt_cell = ws.cell(row=row, column=7, value=float(quotation.grand_total))
    gt_cell.font = grand_total_font
    gt_cell.number_format = '#,##0.00'
    gt_cell.border = thin_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    logger.info("Quotation XLSX exported: id=%d", quotation_id)
    return output


def export_quotation_html(db: Session, user_id: int, quotation_id: int) -> str:
    """Generate quotation as HTML (for PDF rendering or preview)."""
    from app.models.user import User
    from app.models.client import Client

    quotation = (
        db.query(Quotation)
        .options(joinedload(Quotation.sections).joinedload(QuotationSection.line_items))
        .join(Project)
        .filter(Quotation.id == quotation_id, Project.user_id == user_id)
        .first()
    )
    if not quotation:
        raise NotFoundError("Quotation")

    project = db.query(Project).filter(Project.id == quotation.project_id).first()
    designer = db.query(User).filter(User.id == user_id).first()
    client = db.query(Client).filter(Client.id == project.client_id).first() if project else None

    sections_html = ""
    serial = 1
    for section in sorted(quotation.sections, key=lambda s: s.sort_order):
        items_html = ""
        for item in sorted(section.line_items, key=lambda i: i.sort_order):
            items_html += f"""<tr>
                <td style="padding:6px;border:1px solid #ddd;text-align:center">{serial}</td>
                <td style="padding:6px;border:1px solid #ddd">{item.room_name or '-'}</td>
                <td style="padding:6px;border:1px solid #ddd">{item.item_name}</td>
                <td style="padding:6px;border:1px solid #ddd;text-align:right">{float(item.quantity):,.2f}</td>
                <td style="padding:6px;border:1px solid #ddd;text-align:center">{item.unit}</td>
                <td style="padding:6px;border:1px solid #ddd;text-align:right">{float(item.rate):,.2f}</td>
                <td style="padding:6px;border:1px solid #ddd;text-align:right;font-weight:500">{float(item.amount):,.2f}</td>
            </tr>"""
            serial += 1

        sections_html += f"""
        <tr style="background:#e8f0fe">
            <td colspan="6" style="padding:8px;border:1px solid #ddd;font-weight:bold;color:#1a56db">{section.section_name}</td>
            <td style="padding:8px;border:1px solid #ddd;text-align:right;font-weight:bold;color:#1a56db">{float(section.section_total):,.2f}</td>
        </tr>
        {items_html}"""

    # Build company header
    logo_html = ""
    if designer and designer.company_logo_url:
        from app.config import settings
        logo_html = f'<img src="{settings.API_URL}{designer.company_logo_url}" style="height:60px;object-fit:contain;margin-right:16px" />'

    company_name = designer.company_name if designer and designer.company_name else ""
    company_address = designer.company_address if designer and designer.company_address else ""
    company_phone = designer.phone if designer and designer.phone else ""
    company_email = designer.email if designer else ""

    client_name = client.name if client else ""
    client_phone = client.phone if client else ""
    client_address = ""
    if client:
        parts = [client.address, client.city, client.state]
        client_address = ", ".join(p for p in parts if p)

    # Contact line: phone + email
    contact_parts = []
    if company_phone:
        contact_parts.append(f"Ph: {company_phone}")
    if company_email:
        contact_parts.append(f"Email: {company_email}")
    contact_line = " | ".join(contact_parts)

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><style>
body {{ font-family: 'Segoe UI', Arial, sans-serif; color: #333; margin: 40px; }}
h1 {{ color: #1a56db; margin-bottom: 4px; }}
table {{ width: 100%; border-collapse: collapse; margin: 20px 0; font-size: 13px; }}
th {{ background: #1a56db; color: white; padding: 8px; border: 1px solid #1a56db; text-align: center; }}
.totals {{ width: 350px; margin-left: auto; margin-top: 20px; }}
.totals td {{ padding: 6px 10px; }}
.grand {{ font-size: 16px; font-weight: bold; color: #1a56db; border-top: 2px solid #1a56db; }}
.company-header {{ text-align: center; margin-bottom: 10px; }}
.client-box {{ background: #f8f9fa; border: 1px solid #e5e7eb; border-radius: 8px; padding: 12px 16px; margin: 16px 0; }}
</style></head><body>
<!-- Company Header — Centered -->
<div class="company-header">
{f'<img src="{settings.API_URL}{designer.company_logo_url}" style="height:70px;object-fit:contain;margin-bottom:8px" /><br/>' if designer and designer.company_logo_url else ''}
<h1 style="margin:0;color:#1a56db;font-size:24px">{company_name or 'Interior Design Studio'}</h1>
{f'<p style="margin:4px 0 0;color:#555;font-size:13px">{company_address}</p>' if company_address else ''}
{f'<p style="margin:3px 0 0;color:#666;font-size:12px">{contact_line}</p>' if contact_line else ''}
</div>
<hr style="border:none;border-top:2px solid #1a56db;margin:12px 0 20px" />

<!-- Client Info -->
<div class="client-box">
<p style="margin:0;font-size:11px;color:#999;text-transform:uppercase;letter-spacing:1px">Prepared For</p>
<p style="margin:4px 0 0;font-size:15px;font-weight:bold;color:#333">{client_name}</p>
{f'<p style="margin:2px 0 0;font-size:13px;color:#666">{client_address}</p>' if client_address else ''}
{f'<p style="margin:2px 0 0;font-size:13px;color:#666">Ph: {client_phone}</p>' if client_phone else ''}
</div>

<h1>QUOTATION</h1>
<p style="color:#666">{project.name if project else 'Project'} | Version {quotation.version} | {quotation.created_at.strftime('%d %B %Y')}</p>
<table>
<thead><tr>
    <th>S.No</th><th>Room</th><th>Item</th><th>Qty</th><th>Unit</th><th>Rate (₹)</th><th>Amount (₹)</th>
</tr></thead>
<tbody>{sections_html}</tbody>
</table>
<table class="totals">
    <tr><td style="text-align:right">Subtotal</td><td style="text-align:right;font-weight:500">₹{float(quotation.subtotal):,.2f}</td></tr>
    <tr><td style="text-align:right">GST ({quotation.tax_rate}%)</td><td style="text-align:right;font-weight:500">₹{float(quotation.tax_amount):,.2f}</td></tr>
    <tr style="border-top:1px solid #ddd"><td style="text-align:right;padding-top:6px">Total (incl. GST)</td><td style="text-align:right;font-weight:600;padding-top:6px">₹{float(quotation.subtotal + quotation.tax_amount):,.2f}</td></tr>
    {"<tr><td style='text-align:right;color:#059669'>Special Client Discount (" + f"{float(quotation.discount_value)}%)</td><td style='text-align:right;color:#059669;font-weight:600'>-₹{float(quotation.discount_amount):,.2f}</td></tr>" if quotation.discount_amount and quotation.discount_type and quotation.discount_type.value == "percentage" else ("" if not quotation.discount_amount else "<tr><td style='text-align:right;color:#059669'>Discount</td><td style='text-align:right;color:#059669;font-weight:600'>-₹" + f"{float(quotation.discount_amount):,.2f}</td></tr>")}
    <tr class="grand"><td style="text-align:right;padding-top:10px">GRAND TOTAL</td><td style="text-align:right;padding-top:10px">₹{float(quotation.grand_total):,.2f}</td></tr>
</table>
{f'<p style="margin-top:30px;font-size:12px;color:#666"><strong>Notes:</strong> {quotation.notes}</p>' if quotation.notes else ''}
{f'<p style="font-size:12px;color:#666"><strong>Terms:</strong> {quotation.terms_and_conditions}</p>' if quotation.terms_and_conditions else ''}
<p style="margin-top:40px;text-align:center;color:#999;font-size:11px">Generated by DesignBid — Interior Design Proposals</p>
</body></html>"""

    return html
